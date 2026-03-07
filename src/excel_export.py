from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

STATUS_FILLS = {
    "matched": PatternFill(fill_type="solid", fgColor="E2F0D9"),
    "review": PatternFill(fill_type="solid", fgColor="FFF2CC"),
    "unmatched": PatternFill(fill_type="solid", fgColor="FCE4D6"),
}


def autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            try:
                value_length = len(str(cell.value)) if cell.value is not None else 0
                if value_length > max_length:
                    max_length = value_length
            except Exception:
                pass

        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)


def format_worksheet(ws, freeze_cell: str = "A2") -> None:
    ws.freeze_panes = freeze_cell

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")

    autosize_columns(ws)


def color_status_rows(ws, status_col_name: str = "status") -> None:
    headers = [cell.value for cell in ws[1]]
    if status_col_name not in headers:
        return

    status_col_idx = headers.index(status_col_name) + 1

    for row in range(2, ws.max_row + 1):
        status_value = ws.cell(row=row, column=status_col_idx).value
        if status_value in STATUS_FILLS:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = STATUS_FILLS[status_value]


def style_summary_sheet(ws) -> None:
    for row in range(2, ws.max_row + 1):
        metric_cell = ws.cell(row=row, column=1)
        value_cell = ws.cell(row=row, column=2)

        metric_cell.font = Font(bold=True)
        value_cell.alignment = Alignment(horizontal="right")


def export_to_excel(
    summary_df: pd.DataFrame,
    matched_df: pd.DataFrame,
    review_df: pd.DataFrame,
    unmatched_bank_df: pd.DataFrame,
    unmatched_book_df: pd.DataFrame,
    top_review_df: pd.DataFrame,
    top_unmatched_bank_df: pd.DataFrame,
    top_unmatched_book_df: pd.DataFrame,
    output_path: Path,
) -> str:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Resumen", index=False)

        matched_df.to_excel(writer, sheet_name="Matched", index=False)
        review_df.to_excel(writer, sheet_name="Review", index=False)
        unmatched_bank_df.to_excel(writer, sheet_name="Unmatched_Banco", index=False)
        unmatched_book_df.to_excel(writer, sheet_name="Unmatched_Libro", index=False)

        top_review_df.to_excel(writer, sheet_name="Top_Review", index=False)
        top_unmatched_bank_df.to_excel(writer, sheet_name="Top_Unmatched_Banco", index=False)
        top_unmatched_book_df.to_excel(writer, sheet_name="Top_Unmatched_Libro", index=False)

    wb = load_workbook(output_path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        format_worksheet(ws)

        if sheet_name in ["Matched", "Review"]:
            color_status_rows(ws, status_col_name="status")

    style_summary_sheet(wb["Resumen"])

    wb.save(output_path)
    return str(output_path)